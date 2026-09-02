
from flask import Flask, render_template, request, redirect, url_for, session,send_file, flash
from flask_bcrypt import Bcrypt

from ai.demand_prediction import (
    predict_food_demand,
    get_demand_level,
    get_location_demand,
    predict_location_demand,
    generate_recommendations
)

from database import get_connection
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)


app = Flask(__name__)

app.secret_key = "foodlink-development-secret-key"

bcrypt = Bcrypt(app)


# ============================================================
# BUILT-IN ADMINISTRATOR ACCOUNT
#
# Administrators are provisioned by the system, not through
# public registration — the registration form only ever creates
# Student or Donor accounts. These are the credentials that
# control admin access; they are checked directly in the login
# route below, not looked up from something the user submitted
# on the register form.
# ============================================================

ADMIN_EMAIL = "admin@foodlink.com"
ADMIN_PASSWORD = "admin1234"


def ensure_admin_account():
    """Makes sure a real row exists in the users table for the
    built-in administrator, so admin actions — which log against
    a user_id — always have a valid user to reference. Access is
    still controlled by the ADMIN_EMAIL / ADMIN_PASSWORD constants
    above, not by this row's stored password."""

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT user_id, role, account_status
        FROM users
        WHERE email = %s
        """,
        (ADMIN_EMAIL,)
    )

    admin_row = cursor.fetchone()

    if admin_row is None:

        hashed_password = bcrypt.generate_password_hash(
            ADMIN_PASSWORD
        ).decode("utf-8")

        cursor.execute(
            """
            INSERT INTO users
            (name, email, password, role, account_status)
            VALUES (%s, %s, %s, 'Admin', 'Approved')
            """,
            ("Administrator", ADMIN_EMAIL, hashed_password)
        )

        connection.commit()

        admin_id = cursor.lastrowid

    else:

        admin_id = admin_row["user_id"]

        # Self-heal if this email was ever created some other way
        if (
            admin_row["role"] != "Admin"
            or admin_row["account_status"] != "Approved"
        ):

            cursor.execute(
                """
                UPDATE users
                SET role = 'Admin', account_status = 'Approved'
                WHERE user_id = %s
                """,
                (admin_id,)
            )

            connection.commit()

    cursor.close()
    connection.close()

    return admin_id


# ============================================================
# AUDIT LOGGING
# ============================================================

def log_activity(action, description, user_id=None):

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        INSERT INTO activity_logs
        (
            user_id,
            action,
            description
        )
        VALUES
        (
            %s,
            %s,
            %s
        )
        """,
        (
            user_id,
            action,
            description
        )
    )

    connection.commit()

    cursor.close()
    connection.close()


# ============================================================
# NOTIFICATIONS
# ============================================================

def create_notification(user_id, message):

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        INSERT INTO notifications
        (user_id, message)
        VALUES (%s, %s)
        """,
        (user_id, message)
    )

    connection.commit()

    cursor.close()
    connection.close()


# ============================================================
# FOOD MATCHING
# ============================================================

def find_matches():

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Get all pending food requests
    cursor.execute(
        """
        SELECT *
        FROM food_requests
        WHERE status = 'Pending'
        """
    )

    requests = cursor.fetchall()

    for food_request in requests:

        # Get available donations that have not expired
        cursor.execute(
            """
            SELECT *
            FROM food_donations
            WHERE status = 'Available'
            AND expiry_date >= CURDATE()
            """
        )

        donations = cursor.fetchall()

        for donation in donations:

            # -----------------------------------------
            # FOOD MATCHING
            # -----------------------------------------

            requested_food = (
                food_request["description"]
                .lower()
                .strip()
            )

            donated_food = (
                donation["food_name"]
                .lower()
                .strip()
            )

            food_match = (
                donated_food in requested_food
                or requested_food in donated_food
            )

            if not food_match:
                continue

            # -----------------------------------------
            # LOCATION MATCHING
            # -----------------------------------------

            request_location = (
                food_request["location"]
                .lower()
                .strip()
            )

            donation_location = (
                donation["location"]
                .lower()
                .strip()
            )

            # Exact location
            if request_location == donation_location:

                location_score = 2

            # One location contains the other
            elif (
                request_location in donation_location
                or donation_location in request_location
            ):

                location_score = 1

            # Different locations
            else:

                location_score = 0

            # -----------------------------------------
            # MATCH SCORE
            # -----------------------------------------

            # Food must match.
            # Location determines the priority.

            if location_score == 2:

                match_priority = "High"

            elif location_score == 1:

                match_priority = "Medium"

            else:

                match_priority = "Low"

            # -----------------------------------------
            # CHECK EXISTING MATCH
            # -----------------------------------------

            cursor.execute(
                """
                SELECT match_id
                FROM food_matches
                WHERE request_id = %s
                AND donation_id = %s
                """,
                (
                    food_request["request_id"],
                    donation["donation_id"]
                )
            )

            existing_match = cursor.fetchone()

            # -----------------------------------------
            # CREATE MATCH
            # -----------------------------------------

            if not existing_match:

                cursor.execute(
                    """
                    INSERT INTO food_matches
                    (
                        request_id,
                        donation_id,
                        status
                    )
                    VALUES
                    (
                        %s,
                        %s,
                        'Suggested'
                    )
                    """,
                    (
                        food_request["request_id"],
                        donation["donation_id"]
                    )
                )

    connection.commit()

    cursor.close()
    connection.close()


# ============================================================
# HOME
# ============================================================

@app.route("/")
def home():

    return redirect(url_for("login"))


# ============================================================
# TEST DATABASE
# ============================================================

@app.route("/test-db")
def test_database():

    connection = get_connection()

    if connection.is_connected():

        connection.close()

        return "Database connection successful!"

    return "Database connection failed."


# ============================================================
# REGISTER
# ============================================================

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]
        role = request.form["role"]

        # Public registration only ever creates Student or Donor
        # accounts — administrators are provisioned separately.
        if role not in ("Student", "Donor"):

            return "Invalid account type."

        connection = get_connection()
        cursor = connection.cursor()

        # Check whether email already exists
        cursor.execute(
            """
            SELECT user_id
            FROM users
            WHERE email = %s
            """,
            (email,)
        )

        existing_user = cursor.fetchone()

        if existing_user:

            cursor.close()
            connection.close()

            return "An account with this email already exists."

        # Hash password
        hashed_password = bcrypt.generate_password_hash(
            password
        ).decode("utf-8")

        sql = """
            INSERT INTO users
            (
                name,
                email,
                password,
                role
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """

        cursor.execute(
            sql,
            (
                name,
                email,
                hashed_password,
                role
            )
        )

        connection.commit()

        cursor.close()
        connection.close()

        flash(
            "Account created! An administrator will review it "
            "before you can log in.",
            "success"
        )

        return redirect(
            url_for("login")
        )

    return render_template("auth.html", mode="register")


# ============================================================
# LOGIN
# ============================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        # Built-in administrator login — checked against the
        # hardcoded credentials, not looked up as a normal user.
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:

            admin_id = ensure_admin_account()

            session["user_id"] = admin_id
            session["name"] = "Administrator"
            session["role"] = "Admin"

            return redirect(
                url_for("admin_dashboard")
            )

        connection = get_connection()
        cursor = connection.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT *
            FROM users
            WHERE email = %s
            """,
            (email,)
        )

        user = cursor.fetchone()

        cursor.close()
        connection.close()

        if user is None:

            return "Invalid email or password."

        password_correct = bcrypt.check_password_hash(
            user["password"],
            password
        )

        if not password_correct:

            return "Invalid email or password."

        # Check account status
        if user["account_status"] == "Rejected":

            return (
                "Your account has been rejected "
                "by the administrator."
            )

        if user["account_status"] != "Approved":

            return (
                "Your account is waiting "
                "for administrator approval."
            )

        # Store user information in session
        session["user_id"] = user["user_id"]
        session["name"] = user["name"]
        session["role"] = user["role"]

        # Dashboard redirect
        if user["role"] == "Student":

            return redirect(
                url_for("student_dashboard")
            )

        elif user["role"] == "Donor":

            return redirect(
                url_for("donor_dashboard")
            )

        elif user["role"] == "Admin":

            return redirect(
                url_for("admin_dashboard")
            )

        return "Unknown user role."

    return render_template("auth.html", mode="login")


# ============================================================
# STUDENT DASHBOARD
# ============================================================

@app.route("/student/dashboard")
def student_dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # ========================================================
    # STUDENT REQUESTS
    # ========================================================

    cursor.execute(
        """
        SELECT
            request_id,
            urgency,
            location,
            description,
            request_date,
            status
        FROM food_requests
        WHERE student_id = %s
        ORDER BY request_date DESC
        """,
        (session["user_id"],)
    )

    requests = cursor.fetchall()

    # ========================================================
    # REQUEST STATISTICS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE student_id = %s
        """,
        (session["user_id"],)
    )

    total_requests = cursor.fetchone()["total"]

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE student_id = %s
        AND status = 'Pending'
        """,
        (session["user_id"],)
    )

    pending_requests = cursor.fetchone()["total"]

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE student_id = %s
        AND status = 'Approved'
        """,
        (session["user_id"],)
    )

    approved_requests = cursor.fetchone()["total"]

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE student_id = %s
        AND status = 'Completed'
        """,
        (session["user_id"],)
    )

    completed_requests = cursor.fetchone()["total"]

    # ========================================================
    # AVAILABLE MATCHES
    # ========================================================

    find_matches()

    cursor.execute(
        """
        SELECT
            fm.match_id,
            fm.status AS match_status,
            fd.food_name,
            fd.quantity,
            fd.location,
            fd.expiry_date,
            fd.pickup_time
        FROM food_matches fm

        JOIN food_requests fr
            ON fm.request_id = fr.request_id

        JOIN food_donations fd
            ON fm.donation_id = fd.donation_id

        WHERE fr.student_id = %s

        AND fd.status IN (
            'Available',
            'Reserved'
        )

        ORDER BY fm.match_date DESC
        LIMIT 5
        """,
        (session["user_id"],)
    )

    matches = cursor.fetchall()

    # ========================================================
    # RECENT NOTIFICATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT
            notification_id,
            message,
            is_read,
            created_at
        FROM notifications
        WHERE user_id = %s
        ORDER BY created_at DESC
        LIMIT 5
        """,
        (session["user_id"],)
    )

    notifications = cursor.fetchall()

    # ========================================================
    # UNREAD NOTIFICATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM notifications
        WHERE user_id = %s
        AND is_read = 0
        """,
        (session["user_id"],)
    )

    unread_notifications = cursor.fetchone()["total"]

    cursor.close()
    connection.close()

    statistics = {
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "approved_requests": approved_requests,
        "completed_requests": completed_requests
    }

    return render_template(
        "student/dashboard.html",
        name=session["name"],
        requests=requests,
        matches=matches,
        notifications=notifications,
        unread_notifications=unread_notifications,
        statistics=statistics
    )


# ============================================================
# STUDENT REQUEST FOOD
# ============================================================

@app.route("/student/request-food", methods=["GET", "POST"])
def request_food():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    if request.method == "POST":

        urgency = request.form["urgency"]
        location = request.form["location"]
        description = request.form["description"]

        connection = get_connection()
        cursor = connection.cursor()

        sql = """
            INSERT INTO food_requests
            (
                student_id,
                urgency,
                location,
                description
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """

        cursor.execute(
            sql,
            (
                session["user_id"],
                urgency,
                location,
                description
            )
        )

        connection.commit()

        cursor.close()
        connection.close()

        # AUDIT LOG
        log_activity(
            "Food Request Submitted",
            "Student submitted a new food assistance request.",
            session["user_id"]
        )

        # Notify student
        create_notification(
            session["user_id"],
            "Your food assistance request has been submitted successfully."
        )

        return redirect(
            url_for("student_dashboard")
        )

    return render_template(
        "student/request_food.html"
    )


# ============================================================
# DONOR DASHBOARD
# ============================================================

@app.route("/donor/dashboard")
def donor_dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Donor":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            donation_id,
            food_name,
            quantity,
            expiry_date,
            location,
            pickup_time,
            status
        FROM food_donations
        WHERE donor_id = %s
        ORDER BY donation_id DESC
        """,
        (session["user_id"],)
    )

    donations = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "donor/dashboard.html",
        name=session["name"],
        donations=donations
    )


# ============================================================
# DONOR ADD DONATION
# ============================================================

@app.route("/donor/add-donation", methods=["GET", "POST"])
def add_donation():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Donor":
        return "Access denied."

    if request.method == "POST":

        food_name = request.form["food_name"]
        quantity = request.form["quantity"]
        expiry_date = request.form["expiry_date"]
        location = request.form["location"]
        pickup_time = request.form["pickup_time"]

        connection = get_connection()
        cursor = connection.cursor()

        sql = """
            INSERT INTO food_donations
            (
                donor_id,
                food_name,
                quantity,
                expiry_date,
                location,
                pickup_time
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
        """

        cursor.execute(
            sql,
            (
                session["user_id"],
                food_name,
                quantity,
                expiry_date,
                location,
                pickup_time
            )
        )

        connection.commit()

        cursor.close()
        connection.close()

        # AUDIT LOG
        log_activity(
            "Donation Added",
            f"Donation '{food_name}' with quantity '{quantity}' was added.",
            session["user_id"]
        )

        return redirect(
            url_for("donor_dashboard")
        )

    return render_template(
        "donor/add_donation.html"
    )


# ============================================================
# STUDENT MATCHES
# ============================================================

@app.route("/student/matches")
def student_matches():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    find_matches()

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            fm.match_id,
            fm.status AS match_status,
            fd.food_name,
            fd.quantity,
            fd.location,
            fd.expiry_date,
            fd.pickup_time
        FROM food_matches fm

        JOIN food_requests fr
            ON fm.request_id = fr.request_id

        JOIN food_donations fd
            ON fm.donation_id = fd.donation_id

        WHERE fr.student_id = %s

        AND fd.status IN (
            'Available',
            'Reserved'
        )

        ORDER BY fm.match_date DESC
        """,
        (session["user_id"],)
    )

    matches = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "student/matches.html",
        matches=matches
    )


# ============================================================
# STUDENT ACCEPT DONATION
# ============================================================

@app.route(
    "/student/accept-donation/<int:match_id>",
    methods=["POST"]
)
def accept_donation(match_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Find match belonging to this student
    cursor.execute(
        """
        SELECT
            fm.match_id,
            fm.request_id,
            fm.donation_id,
            fm.status AS match_status,
            fd.donor_id,
            fd.status AS donation_status
        FROM food_matches fm

        JOIN food_requests fr
            ON fm.request_id = fr.request_id

        JOIN food_donations fd
            ON fm.donation_id = fd.donation_id

        WHERE fm.match_id = %s
        AND fr.student_id = %s
        """,
        (
            match_id,
            session["user_id"]
        )
    )

    match = cursor.fetchone()

    if match is None:

        cursor.close()
        connection.close()

        return "Match not found."

    if match["match_status"] != "Suggested":

        cursor.close()
        connection.close()

        return "This donation has already been processed."

    if match["donation_status"] != "Available":

        cursor.close()
        connection.close()

        return "This donation is no longer available."

    # Accept match
    cursor.execute(
        """
        UPDATE food_matches
        SET status = 'Accepted'
        WHERE match_id = %s
        """,
        (match_id,)
    )

    # Reserve donation
    cursor.execute(
        """
        UPDATE food_donations
        SET status = 'Reserved'
        WHERE donation_id = %s
        """,
        (match["donation_id"],)
    )

    # Approve request
    cursor.execute(
        """
        UPDATE food_requests
        SET status = 'Approved'
        WHERE request_id = %s
        """,
        (match["request_id"],)
    )

    # Reject other suggested matches
    cursor.execute(
        """
        UPDATE food_matches
        SET status = 'Rejected'
        WHERE donation_id = %s
        AND match_id != %s
        AND status = 'Suggested'
        """,
        (
            match["donation_id"],
            match_id
        )
    )

    # Notify donor
    cursor.execute(
        """
        INSERT INTO notifications
        (
            user_id,
            message
        )
        VALUES
        (
            %s,
            %s
        )
        """,
        (
            match["donor_id"],
            "Your food donation has been accepted by a student."
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Donation Accepted",
        f"Student accepted donation match #{match_id}.",
        session["user_id"]
    )

    return redirect(
        url_for("student_matches")
    )


# ============================================================
# ADMIN DASHBOARD
# ============================================================

@app.route("/admin/dashboard")
def admin_dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    # ========================================================
    # DATABASE CONNECTION
    # ========================================================

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # ========================================================
    # STUDENTS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'Student'
        """
    )

    students = cursor.fetchone()["total"]

    # ========================================================
    # DONORS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'Donor'
        """
    )

    donors = cursor.fetchone()["total"]

    # ========================================================
    # PENDING REQUESTS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE status = 'Pending'
        """
    )

    pending_requests = cursor.fetchone()["total"]

    # ========================================================
    # AVAILABLE DONATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_donations
        WHERE status = 'Available'
        """
    )

    available_donations = cursor.fetchone()["total"]

    # ========================================================
    # RESERVED DONATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_donations
        WHERE status = 'Reserved'
        """
    )

    reserved_donations = cursor.fetchone()["total"]

    # ========================================================
    # COLLECTED DONATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_donations
        WHERE status = 'Collected'
        """
    )

    collected_donations = cursor.fetchone()["total"]

    # ========================================================
    # COMPLETED REQUESTS
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE status = 'Completed'
        """
    )

    completed_requests = cursor.fetchone()["total"]

    # ========================================================
    # ACCEPTED MATCHES
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_matches
        WHERE status = 'Accepted'
        """
    )

    accepted_matches = cursor.fetchone()["total"]

    # ========================================================
    # COMPLETED MATCHES
    # ========================================================

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_matches
        WHERE status = 'Completed'
        """
    )

    completed_matches = cursor.fetchone()["total"]

    # ========================================================
    # RECENT REQUESTS
    # ========================================================

    cursor.execute(
        """
        SELECT
            fr.request_id,
            fr.description,
            fr.urgency,
            fr.location,
            fr.status,
            fr.request_date,
            u.name AS student_name

        FROM food_requests fr

        JOIN users u
            ON fr.student_id = u.user_id

        ORDER BY fr.request_date DESC

        LIMIT 5
        """
    )

    recent_requests = cursor.fetchall()

    # ========================================================
    # RECENT DONATIONS
    # ========================================================

    cursor.execute(
        """
        SELECT
            fd.donation_id,
            fd.food_name,
            fd.quantity,
            fd.location,
            fd.status,
            fd.expiry_date,
            u.name AS donor_name

        FROM food_donations fd

        JOIN users u
            ON fd.donor_id = u.user_id

        ORDER BY fd.donation_id DESC

        LIMIT 5
        """
    )

    recent_donations = cursor.fetchall()

    # ========================================================
    # CHARTS & STATISTICS
    # ========================================================

    # --------------------------------------------------------
    # REQUESTS BY STATUS
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT
            status,
            COUNT(*) AS count

        FROM food_requests

        GROUP BY status
        """
    )

    request_status_data = cursor.fetchall()

    # --------------------------------------------------------
    # DONATIONS BY STATUS
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT
            status,
            COUNT(*) AS count

        FROM food_donations

        GROUP BY status
        """
    )

    donation_status_data = cursor.fetchall()

    # --------------------------------------------------------
    # REQUESTS BY LOCATION
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT
            location,
            COUNT(*) AS count

        FROM food_requests

        GROUP BY location

        ORDER BY count DESC
        """
    )

    request_location_data = cursor.fetchall()

    # --------------------------------------------------------
    # DONATIONS BY LOCATION
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT
            location,
            COUNT(*) AS count

        FROM food_donations

        GROUP BY location

        ORDER BY count DESC
        """
    )

    donation_location_data = cursor.fetchall()

    # --------------------------------------------------------
    # TOTAL REQUESTS
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_requests
        """
    )

    total_requests = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # TOTAL DONATIONS
    # --------------------------------------------------------

    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM food_donations
        """
    )

    total_donations = cursor.fetchone()["total"]

    # ========================================================
    # CLOSE DATABASE
    # ========================================================

    cursor.close()
    connection.close()

    # ========================================================
    # AI FOOD DEMAND PREDICTION
    # ========================================================

    prediction = predict_food_demand()

    demand_level = get_demand_level(
        prediction
    )

    # ========================================================
    # CURRENT DEMAND BY LOCATION
    # ========================================================

    location_demand = get_location_demand()

    # ========================================================
    # AI LOCATION PREDICTIONS
    # ========================================================

    location_predictions = predict_location_demand()

    # ========================================================
    # AI RECOMMENDATIONS
    # ========================================================

    recommendations = generate_recommendations()

    # ========================================================
    # STATISTICS OBJECT
    # ========================================================

    statistics = {

        "students": students,

        "donors": donors,

        "pending_requests": pending_requests,

        "available_donations": available_donations,

        "reserved_donations": reserved_donations,

        "collected_donations": collected_donations,

        "completed_requests": completed_requests,

        "accepted_matches": accepted_matches,

        "completed_matches": completed_matches
    }

    # ========================================================
    # RENDER DASHBOARD
    # ========================================================

    return render_template(

        "admin/dashboard.html",

        # Existing dashboard data
        name=session["name"],

        statistics=statistics,

        prediction=prediction,

        demand_level=demand_level,

        location_demand=location_demand,

        location_predictions=location_predictions,

        recommendations=recommendations,

        recent_requests=recent_requests,

        recent_donations=recent_donations,

        # ====================================================
        # CHART DATA
        # ====================================================

        total_requests=total_requests,

        total_donations=total_donations,

        completed_requests=completed_requests,

        pending_requests=pending_requests,

        request_status_data=request_status_data,

        donation_status_data=donation_status_data,

        request_location_data=request_location_data,

        donation_location_data=donation_location_data
    )



# ============================================================
# ADMIN USERS
# ============================================================

@app.route("/admin/users")
def admin_users():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            user_id,
            name,
            email,
            role,
            account_status
        FROM users
        ORDER BY user_id DESC
        """
    )

    users = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "admin/users.html",
        users=users
    )


# ============================================================
# ADMIN APPROVE USER
# ============================================================

@app.route(
    "/admin/approve-user/<int:user_id>",
    methods=["POST"]
)
def approve_user(user_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        UPDATE users
        SET account_status = 'Approved'
        WHERE user_id = %s
        """,
        (user_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "User Approved",
        f"User #{user_id} was approved by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_users")
    )


# ============================================================
# ADMIN REJECT USER
# ============================================================

@app.route(
    "/admin/reject-user/<int:user_id>",
    methods=["POST"]
)
def reject_user(user_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        UPDATE users
        SET account_status = 'Rejected'
        WHERE user_id = %s
        """,
        (user_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "User Rejected",
        f"User #{user_id} was rejected by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_users")
    )


# ============================================================
# ADMIN FOOD REQUESTS
# ============================================================

@app.route("/admin/requests")
def admin_requests():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            fr.request_id,
            fr.description,
            fr.urgency,
            fr.location,
            fr.status,
            fr.request_date,
            u.name AS student_name
        FROM food_requests fr

        JOIN users u
            ON fr.student_id = u.user_id

        ORDER BY fr.request_date DESC
        """
    )

    requests = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "admin/requests.html",
        requests=requests
    )


# ============================================================
# ADMIN DONATIONS
# ============================================================

@app.route("/admin/donations", methods=["GET"])
def admin_donations():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip()
    location_filter = request.args.get("location", "").strip()

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    sql = """
        SELECT
            d.donation_id,
            d.food_name,
            d.quantity,
            d.expiry_date,
            d.location,
            d.pickup_time,
            d.status,
            u.name AS donor_name,
            u.email AS donor_email
        FROM food_donations d
        JOIN users u
            ON d.donor_id = u.user_id
        WHERE 1=1
    """

    params = []

    if search:

        sql += """
            AND (
                d.food_name LIKE %s
                OR u.name LIKE %s
                OR u.email LIKE %s
            )
        """

        search_value = "%" + search + "%"

        params.extend([
            search_value,
            search_value,
            search_value
        ])

    if status_filter:

        sql += """
            AND d.status = %s
        """

        params.append(status_filter)

    if location_filter:

        sql += """
            AND d.location LIKE %s
        """

        params.append("%" + location_filter + "%")

    sql += """
        ORDER BY d.donation_id DESC
    """

    cursor.execute(
        sql,
        tuple(params)
    )

    donations = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "admin/donations.html",
        donations=donations,
        search=search,
        status_filter=status_filter,
        location_filter=location_filter
    )


# ============================================================
# ADMIN COLLECT DONATION
# ============================================================

@app.route(
    "/admin/collect-donation/<int:donation_id>",
    methods=["POST"]
)
def collect_donation(donation_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Find donation
    cursor.execute(
        """
        SELECT
            donation_id,
            status
        FROM food_donations
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    donation = cursor.fetchone()

    if donation is None:

        cursor.close()
        connection.close()

        return "Donation not found."

    if donation["status"] != "Reserved":

        cursor.close()
        connection.close()

        return "This donation is not ready to be collected."

    # Mark donation as collected
    cursor.execute(
        """
        UPDATE food_donations
        SET status = 'Collected'
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    # Find accepted match
    cursor.execute(
        """
        SELECT
            fm.match_id,
            fm.request_id,
            fr.student_id
        FROM food_matches fm

        JOIN food_requests fr
            ON fm.request_id = fr.request_id

        WHERE fm.donation_id = %s
        AND fm.status = 'Accepted'
        """,
        (donation_id,)
    )

    match = cursor.fetchone()

    if match:

        # Complete match
        cursor.execute(
            """
            UPDATE food_matches
            SET status = 'Completed'
            WHERE match_id = %s
            """,
            (match["match_id"],)
        )

        # Complete request
        cursor.execute(
            """
            UPDATE food_requests
            SET status = 'Completed'
            WHERE request_id = %s
            """,
            (match["request_id"],)
        )

        # Notify student
        cursor.execute(
            """
            INSERT INTO notifications
            (
                user_id,
                message
            )
            VALUES
            (
                %s,
                %s
            )
            """,
            (
                match["student_id"],
                "Your food donation has been collected and your food assistance request has been completed."
            )
        )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Donation Collected",
        f"Donation #{donation_id} was marked as collected.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_donations")
    )


# ============================================================
# STUDENT NOTIFICATIONS
# ============================================================

@app.route("/student/notifications")
def student_notifications():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            notification_id,
            message,
            is_read,
            created_at
        FROM notifications
        WHERE user_id = %s
        ORDER BY created_at DESC
        """,
        (session["user_id"],)
    )

    notifications = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "student/notifications.html",
        notifications=notifications
    )


# ============================================================
# DONOR NOTIFICATIONS
# ============================================================

@app.route("/donor/notifications")
def donor_notifications():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Donor":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            notification_id,
            message,
            is_read,
            created_at
        FROM notifications
        WHERE user_id = %s
        ORDER BY created_at DESC
        """,
        (session["user_id"],)
    )

    notifications = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "donor/notifications.html",
        notifications=notifications
    )


# ============================================================
# ADMIN EDIT USER
# ============================================================

@app.route(
    "/admin/edit-user/<int:user_id>",
    methods=["GET", "POST"]
)
def edit_user(user_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Get user
    cursor.execute(
        """
        SELECT
            user_id,
            name,
            email,
            role,
            account_status
        FROM users
        WHERE user_id = %s
        """,
        (user_id,)
    )

    user = cursor.fetchone()

    if user is None:

        cursor.close()
        connection.close()

        return "User not found."

    # ========================================================
    # UPDATE USER
    # ========================================================

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        role = request.form["role"]
        account_status = request.form["account_status"]

        cursor.execute(
            """
            SELECT user_id
            FROM users
            WHERE email = %s
            AND user_id != %s
            """,
            (
                email,
                user_id
            )
        )

        existing_user = cursor.fetchone()

        if existing_user:

            cursor.close()
            connection.close()

            return "Another account is already using this email."

        cursor.execute(
            """
            UPDATE users
            SET
                name = %s,
                email = %s,
                role = %s,
                account_status = %s
            WHERE user_id = %s
            """,
            (
                name,
                email,
                role,
                account_status,
                user_id
            )
        )

        connection.commit()

        cursor.close()
        connection.close()

        # AUDIT LOG
        log_activity(
            "User Edited",
            f"User #{user_id} was edited by the administrator.",
            session["user_id"]
        )

        return redirect(
            url_for("admin_users")
        )

    cursor.close()
    connection.close()

    return render_template(
        "admin/edit_user.html",
        user=user
    )


# ============================================================
# ADMIN DELETE USER
# ============================================================

@app.route(
    "/admin/delete-user/<int:user_id>",
    methods=["POST"]
)
def delete_user(user_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    # Prevent administrator from deleting themselves
    if user_id == session["user_id"]:

        return "You cannot delete your own administrator account."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Check user exists
    cursor.execute(
        """
        SELECT
            user_id,
            name,
            role
        FROM users
        WHERE user_id = %s
        """,
        (user_id,)
    )

    user = cursor.fetchone()

    if user is None:

        cursor.close()
        connection.close()

        return "User not found."

    # IMPORTANT:
    # Log before deleting the user because the user
    # will no longer exist afterwards.
    log_description = (
        f"User #{user_id} ({user['name']}) "
        f"was permanently deleted by the administrator."
    )

    # Delete notifications
    cursor.execute(
        """
        DELETE FROM notifications
        WHERE user_id = %s
        """,
        (user_id,)
    )

    # If Student
    if user["role"] == "Student":

        cursor.execute(
            """
            DELETE fm
            FROM food_matches fm
            JOIN food_requests fr
                ON fm.request_id = fr.request_id
            WHERE fr.student_id = %s
            """,
            (user_id,)
        )

        cursor.execute(
            """
            DELETE FROM food_requests
            WHERE student_id = %s
            """,
            (user_id,)
        )

    # If Donor
    elif user["role"] == "Donor":

        cursor.execute(
            """
            DELETE fm
            FROM food_matches fm
            JOIN food_donations fd
                ON fm.donation_id = fd.donation_id
            WHERE fd.donor_id = %s
            """,
            (user_id,)
        )

        cursor.execute(
            """
            DELETE FROM food_donations
            WHERE donor_id = %s
            """,
            (user_id,)
        )

    # Delete user
    cursor.execute(
        """
        DELETE FROM users
        WHERE user_id = %s
        """,
        (user_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "User Deleted",
        log_description,
        session["user_id"]
    )

    return redirect(
        url_for("admin_users")
    )


# ============================================================
# ADMIN DELETE FOOD REQUEST
# ============================================================

@app.route(
    "/admin/delete-request/<int:request_id>",
    methods=["POST"]
)
def delete_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            request_id
        FROM food_requests
        WHERE request_id = %s
        """,
        (request_id,)
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    # Delete related matches first
    cursor.execute(
        """
        DELETE FROM food_matches
        WHERE request_id = %s
        """,
        (request_id,)
    )

    # Delete request
    cursor.execute(
        """
        DELETE FROM food_requests
        WHERE request_id = %s
        """,
        (request_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Request Deleted",
        f"Food request #{request_id} was permanently deleted by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_requests")
    )


# ============================================================
# ADMIN APPROVE FOOD REQUEST
# ============================================================

@app.route(
    "/admin/approve-request/<int:request_id>",
    methods=["POST"]
)
def approve_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            fr.request_id,
            fr.student_id,
            fr.status
        FROM food_requests fr
        WHERE fr.request_id = %s
        """,
        (request_id,)
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    if food_request["status"] != "Pending":

        cursor.close()
        connection.close()

        return "This request has already been processed."

    # Approve request
    cursor.execute(
        """
        UPDATE food_requests
        SET status = 'Approved'
        WHERE request_id = %s
        """,
        (request_id,)
    )

    # Notify student
    cursor.execute(
        """
        INSERT INTO notifications
        (
            user_id,
            message
        )
        VALUES
        (
            %s,
            %s
        )
        """,
        (
            food_request["student_id"],
            "Your food assistance request has been approved by an administrator."
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Request Approved",
        f"Food request #{request_id} was approved by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_requests")
    )


# ============================================================
# ADMIN REJECT FOOD REQUEST
# ============================================================

@app.route(
    "/admin/reject-request/<int:request_id>",
    methods=["POST"]
)
def reject_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            request_id,
            student_id,
            status
        FROM food_requests
        WHERE request_id = %s
        """,
        (request_id,)
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    if food_request["status"] != "Pending":

        cursor.close()
        connection.close()

        return "This request has already been processed."

    # Reject request
    cursor.execute(
        """
        UPDATE food_requests
        SET status = 'Rejected'
        WHERE request_id = %s
        """,
        (request_id,)
    )

    # Notify student
    cursor.execute(
        """
        INSERT INTO notifications
        (
            user_id,
            message
        )
        VALUES
        (
            %s,
            %s
        )
        """,
        (
            food_request["student_id"],
            "Your food assistance request has been rejected by an administrator."
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Request Rejected",
        f"Food request #{request_id} was rejected by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_requests")
    )


# ============================================================
# ADMIN COMPLETE FOOD REQUEST
# ============================================================

@app.route(
    "/admin/complete-request/<int:request_id>",
    methods=["POST"]
)
def complete_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            request_id,
            student_id,
            status
        FROM food_requests
        WHERE request_id = %s
        """,
        (request_id,)
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    if food_request["status"] != "Approved":

        cursor.close()
        connection.close()

        return "Only approved requests can be completed."

    # Mark request as completed
    cursor.execute(
        """
        UPDATE food_requests
        SET status = 'Completed'
        WHERE request_id = %s
        """,
        (request_id,)
    )

    # Notify student
    cursor.execute(
        """
        INSERT INTO notifications
        (
            user_id,
            message
        )
        VALUES
        (
            %s,
            %s
        )
        """,
        (
            food_request["student_id"],
            "Your food assistance request has been completed."
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Request Completed",
        f"Food request #{request_id} was marked as completed by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_requests")
    )


# ============================================================
# ADMIN CANCEL DONATION
# ============================================================

@app.route(
    "/admin/cancel-donation/<int:donation_id>",
    methods=["POST"]
)
def cancel_donation(donation_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            donation_id,
            donor_id,
            status
        FROM food_donations
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    donation = cursor.fetchone()

    if donation is None:

        cursor.close()
        connection.close()

        return "Donation not found."

    if donation["status"] != "Available":

        cursor.close()
        connection.close()

        return "Only available donations can be cancelled."

    # Cancel donation
    cursor.execute(
        """
        UPDATE food_donations
        SET status = 'Cancelled'
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    # Reject suggested matches
    cursor.execute(
        """
        UPDATE food_matches
        SET status = 'Rejected'
        WHERE donation_id = %s
        AND status = 'Suggested'
        """,
        (donation_id,)
    )

    # Notify donor
    cursor.execute(
        """
        INSERT INTO notifications
        (
            user_id,
            message
        )
        VALUES
        (
            %s,
            %s
        )
        """,
        (
            donation["donor_id"],
            "Your food donation has been cancelled by the administrator."
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Donation Cancelled",
        f"Donation #{donation_id} was cancelled by the administrator.",
        session["user_id"]
    )

    return redirect(
        url_for("admin_donations")
    )


# ============================================================
# ADMIN DELETE DONATION
# ============================================================

@app.route(
    "/admin/delete-donation/<int:donation_id>",
    methods=["POST"]
)
def delete_donation(donation_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Check donation
    cursor.execute(
        """
        SELECT
            donation_id,
            food_name,
            quantity,
            status
        FROM food_donations
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    donation = cursor.fetchone()

    if donation is None:

        cursor.close()
        connection.close()

        return "Donation not found."

    if donation["status"] in ("Reserved", "Collected"):

        cursor.close()
        connection.close()

        return (
            "This donation cannot be deleted "
            "because it has already been processed."
        )

    # Save audit information BEFORE deleting
    log_description = (
        f"Donation #{donation_id} "
        f"('{donation['food_name']}', quantity '{donation['quantity']}') "
        f"was permanently deleted by the administrator."
    )

    # Delete related matches first
    cursor.execute(
        """
        DELETE FROM food_matches
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    # Delete donation
    cursor.execute(
        """
        DELETE FROM food_donations
        WHERE donation_id = %s
        """,
        (donation_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Donation Deleted",
        log_description,
        session["user_id"]
    )

    return redirect(
        url_for("admin_donations")
    )


# ============================================================
# STUDENT EDIT FOOD REQUEST
# ============================================================

@app.route(
    "/student/edit-request/<int:request_id>",
    methods=["GET", "POST"]
)
def edit_student_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Get request belonging to logged-in student
    cursor.execute(
        """
        SELECT
            request_id,
            urgency,
            location,
            description,
            status,
            request_date
        FROM food_requests
        WHERE request_id = %s
        AND student_id = %s
        """,
        (
            request_id,
            session["user_id"]
        )
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    if food_request["status"] != "Pending":

        cursor.close()
        connection.close()

        return (
            "This request can no longer be edited "
            "because it has already been processed."
        )

    if request.method == "POST":

        urgency = request.form["urgency"]
        location = request.form["location"]
        description = request.form["description"]

        cursor.execute(
            """
            UPDATE food_requests
            SET
                urgency = %s,
                location = %s,
                description = %s
            WHERE request_id = %s
            AND student_id = %s
            """,
            (
                urgency,
                location,
                description,
                request_id,
                session["user_id"]
            )
        )

        # Remove old suggested matches
        cursor.execute(
            """
            DELETE FROM food_matches
            WHERE request_id = %s
            AND status = 'Suggested'
            """,
            (request_id,)
        )

        connection.commit()

        cursor.close()
        connection.close()

        # AUDIT LOG
        log_activity(
            "Food Request Edited",
            f"Food request #{request_id} was edited by the student.",
            session["user_id"]
        )

        # Notify student
        create_notification(
            session["user_id"],
            "Your food assistance request has been updated successfully."
        )

        return redirect(
            url_for("student_dashboard")
        )

    cursor.close()
    connection.close()

    return render_template(
        "student/edit_request.html",
        food_request=food_request
    )


# ============================================================
# STUDENT CANCEL FOOD REQUEST
# ============================================================

@app.route(
    "/student/cancel-request/<int:request_id>",
    methods=["POST"]
)
def cancel_student_request(request_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Student":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            request_id,
            status
        FROM food_requests
        WHERE request_id = %s
        AND student_id = %s
        """,
        (
            request_id,
            session["user_id"]
        )
    )

    food_request = cursor.fetchone()

    if food_request is None:

        cursor.close()
        connection.close()

        return "Food request not found."

    if food_request["status"] != "Pending":

        cursor.close()
        connection.close()

        return (
            "This request cannot be cancelled "
            "because it has already been processed."
        )

    # Cancel request
    cursor.execute(
        """
        UPDATE food_requests
        SET status = 'Cancelled'
        WHERE request_id = %s
        AND student_id = %s
        """,
        (
            request_id,
            session["user_id"]
        )
    )

    # Reject suggested matches
    cursor.execute(
        """
        UPDATE food_matches
        SET status = 'Rejected'
        WHERE request_id = %s
        AND status = 'Suggested'
        """,
        (request_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Food Request Cancelled",
        f"Food request #{request_id} was cancelled by the student.",
        session["user_id"]
    )

    # Notify student
    create_notification(
        session["user_id"],
        "Your food assistance request has been cancelled."
    )

    return redirect(
        url_for("student_dashboard")
    )


# ============================================================
# DONOR EDIT DONATION
# ============================================================

@app.route(
    "/donor/edit-donation/<int:donation_id>",
    methods=["GET", "POST"]
)
def edit_donation(donation_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Donor":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Get donation belonging to donor
    cursor.execute(
        """
        SELECT
            donation_id,
            food_name,
            quantity,
            expiry_date,
            location,
            pickup_time,
            status
        FROM food_donations
        WHERE donation_id = %s
        AND donor_id = %s
        """,
        (
            donation_id,
            session["user_id"]
        )
    )

    donation = cursor.fetchone()

    if donation is None:

        cursor.close()
        connection.close()

        return "Donation not found."

    if donation["status"] != "Available":

        cursor.close()
        connection.close()

        return (
            "This donation can no longer be edited "
            "because it has already been processed."
        )

    if request.method == "POST":

        food_name = request.form["food_name"]
        quantity = request.form["quantity"]
        expiry_date = request.form["expiry_date"]
        location = request.form["location"]
        pickup_time = request.form["pickup_time"]

        cursor.execute(
            """
            UPDATE food_donations
            SET
                food_name = %s,
                quantity = %s,
                expiry_date = %s,
                location = %s,
                pickup_time = %s
            WHERE donation_id = %s
            AND donor_id = %s
            AND status = 'Available'
            """,
            (
                food_name,
                quantity,
                expiry_date,
                location,
                pickup_time,
                donation_id,
                session["user_id"]
            )
        )

        connection.commit()

        cursor.close()
        connection.close()

        # AUDIT LOG
        log_activity(
            "Donation Edited",
            f"Donation #{donation_id} was edited by the donor.",
            session["user_id"]
        )

        create_notification(
            session["user_id"],
            "Your food donation has been updated successfully."
        )

        return redirect(
            url_for("donor_dashboard")
        )

    cursor.close()
    connection.close()

    return render_template(
        "donor/edit_donation.html",
        donation=donation
    )


# ============================================================
# DONOR CANCEL DONATION
# ============================================================

@app.route(
    "/donor/cancel-donation/<int:donation_id>",
    methods=["POST"]
)
def donor_cancel_donation(donation_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Donor":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # Make sure donation belongs to donor
    cursor.execute(
        """
        SELECT
            donation_id,
            status
        FROM food_donations
        WHERE donation_id = %s
        AND donor_id = %s
        """,
        (
            donation_id,
            session["user_id"]
        )
    )

    donation = cursor.fetchone()

    if donation is None:

        cursor.close()
        connection.close()

        return "Donation not found."

    if donation["status"] != "Available":

        cursor.close()
        connection.close()

        return (
            "This donation can no longer be cancelled "
            "because it has already been processed."
        )

    # Cancel donation
    cursor.execute(
        """
        UPDATE food_donations
        SET status = 'Cancelled'
        WHERE donation_id = %s
        AND donor_id = %s
        AND status = 'Available'
        """,
        (
            donation_id,
            session["user_id"]
        )
    )

    # Reject suggested matches
    cursor.execute(
        """
        UPDATE food_matches
        SET status = 'Rejected'
        WHERE donation_id = %s
        AND status = 'Suggested'
        """,
        (donation_id,)
    )

    connection.commit()

    cursor.close()
    connection.close()

    # AUDIT LOG
    log_activity(
        "Donation Cancelled",
        f"Donation #{donation_id} was cancelled by the donor.",
        session["user_id"]
    )

    create_notification(
        session["user_id"],
        "Your food donation has been cancelled."
    )

    return redirect(
        url_for("donor_dashboard")
    )


# ============================================================
# NOTIFICATIONS
# ============================================================

@app.route("/notifications")
def notifications():

    if "user_id" not in session:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            notification_id,
            message,
            is_read,
            created_at
        FROM notifications
        WHERE user_id = %s
        ORDER BY created_at DESC
        """,
        (session["user_id"],)
    )

    notifications = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "notifications.html",
        notifications=notifications
    )


# ============================================================
# MARK ONE NOTIFICATION AS READ
# ============================================================

@app.route(
    "/notifications/read/<int:notification_id>",
    methods=["POST"]
)
def mark_notification_read(notification_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        UPDATE notifications
        SET is_read = 1
        WHERE notification_id = %s
        AND user_id = %s
        """,
        (
            notification_id,
            session["user_id"]
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    return redirect(
        url_for("notifications")
    )


# ============================================================
# MARK ALL NOTIFICATIONS AS READ
# ============================================================

@app.route(
    "/notifications/read-all",
    methods=["POST"]
)
def mark_all_notifications_read():

    if "user_id" not in session:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        UPDATE notifications
        SET is_read = 1
        WHERE user_id = %s
        """,
        (session["user_id"],)
    )

    connection.commit()

    cursor.close()
    connection.close()

    return redirect(
        url_for("notifications")
    )


# ============================================================
# DELETE NOTIFICATION
# ============================================================

@app.route(
    "/notifications/delete/<int:notification_id>",
    methods=["POST"]
)
def delete_notification(notification_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        DELETE FROM notifications
        WHERE notification_id = %s
        AND user_id = %s
        """,
        (
            notification_id,
            session["user_id"]
        )
    )

    connection.commit()

    cursor.close()
    connection.close()

    return redirect(
        url_for("notifications")
    )


# ============================================================
# ADMIN ACTIVITY / AUDIT LOGS
# ============================================================

@app.route("/admin/activity-logs")
def admin_activity_logs():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(
        """
        SELECT
            a.log_id,
            a.action,
            a.description,
            a.created_at,
            u.name AS user_name,
            u.email AS user_email,
            u.role AS user_role

        FROM activity_logs a

        LEFT JOIN users u
            ON a.user_id = u.user_id

        ORDER BY a.created_at DESC
        """
    )

    logs = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(
        "admin/activity_logs.html",
        logs=logs
    )

# ============================================================
# EXCEL REPORT
# ============================================================

@app.route("/admin/reports/excel")
def admin_excel_report():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # --------------------------------------------------------
    # USERS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            role,
            COUNT(*) AS total
        FROM users
        GROUP BY role
    """)

    user_statistics = cursor.fetchall()

    # --------------------------------------------------------
    # FOOD REQUESTS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            request_id,
            student_id,
            description,
            urgency,
            location,
            status,
            request_date
        FROM food_requests
        ORDER BY request_date DESC
    """)

    requests = cursor.fetchall()

    # --------------------------------------------------------
    # FOOD DONATIONS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            donation_id,
            donor_id,
            food_name,
            quantity,
            expiry_date,
            location,
            pickup_time,
            status
        FROM food_donations
        ORDER BY donation_id DESC
    """)

    donations = cursor.fetchall()

    # --------------------------------------------------------
    # REQUEST STATUS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            status,
            COUNT(*) AS total
        FROM food_requests
        GROUP BY status
    """)

    request_status = cursor.fetchall()

    # --------------------------------------------------------
    # DONATION STATUS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            status,
            COUNT(*) AS total
        FROM food_donations
        GROUP BY status
    """)

    donation_status = cursor.fetchall()

    cursor.close()
    connection.close()

    # ========================================================
    # CREATE EXCEL WORKBOOK
    # ========================================================

    workbook = Workbook()

    # --------------------------------------------------------
    # SUMMARY SHEET
    # --------------------------------------------------------

    summary = workbook.active
    summary.title = "Summary"

    summary["A1"] = "FoodLink Kimberley"
    summary["A2"] = "Administrative Report"
    summary["A3"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    summary["A5"] = "User Statistics"
    summary["A6"] = "Role"
    summary["B6"] = "Total"

    row = 7

    for item in user_statistics:

        summary.cell(
            row=row,
            column=1,
            value=item["role"]
        )

        summary.cell(
            row=row,
            column=2,
            value=item["total"]
        )

        row += 1

    row += 2

    summary.cell(
        row=row,
        column=1,
        value="Request Status"
    )

    row += 1

    summary.cell(
        row=row,
        column=1,
        value="Status"
    )

    summary.cell(
        row=row,
        column=2,
        value="Total"
    )

    row += 1

    for item in request_status:

        summary.cell(
            row=row,
            column=1,
            value=item["status"]
        )

        summary.cell(
            row=row,
            column=2,
            value=item["total"]
        )

        row += 1

    row += 2

    summary.cell(
        row=row,
        column=1,
        value="Donation Status"
    )

    row += 1

    summary.cell(
        row=row,
        column=1,
        value="Status"
    )

    summary.cell(
        row=row,
        column=2,
        value="Total"
    )

    row += 1

    for item in donation_status:

        summary.cell(
            row=row,
            column=1,
            value=item["status"]
        )

        summary.cell(
            row=row,
            column=2,
            value=item["total"]
        )

        row += 1

    # --------------------------------------------------------
    # REQUESTS SHEET
    # --------------------------------------------------------

    request_sheet = workbook.create_sheet(
        "Food Requests"
    )

    request_headers = [
        "Request ID",
        "Student ID",
        "Description",
        "Urgency",
        "Location",
        "Status",
        "Request Date"
    ]

    request_sheet.append(request_headers)

    for item in requests:

        request_sheet.append([
            item["request_id"],
            item["student_id"],
            item["description"],
            item["urgency"],
            item["location"],
            item["status"],
            item["request_date"]
        ])

    # --------------------------------------------------------
    # DONATIONS SHEET
    # --------------------------------------------------------

    donation_sheet = workbook.create_sheet(
        "Food Donations"
    )

    donation_headers = [
        "Donation ID",
        "Donor ID",
        "Food",
        "Quantity",
        "Expiry Date",
        "Location",
        "Pickup Time",
        "Status"
    ]

    donation_sheet.append(donation_headers)

    for item in donations:

        donation_sheet.append([
            item["donation_id"],
            item["donor_id"],
            item["food_name"],
            item["quantity"],
            item["expiry_date"],
            item["location"],
            item["pickup_time"],
            item["status"]
        ])

    # --------------------------------------------------------
    # FORMAT EXCEL
    # --------------------------------------------------------

    for sheet in workbook.worksheets:

        for cell in sheet[1]:

            cell.font = cell.font.copy(
                bold=True
            )

        for column in sheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40
            )

    # --------------------------------------------------------
    # SAVE TO MEMORY
    # --------------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        "FoodLink_Kimberley_Report_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

# ============================================================
# PDF REPORT
# ============================================================

@app.route("/admin/reports/pdf")
def admin_pdf_report():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session["role"] != "Admin":
        return "Access denied."

    connection = get_connection()
    cursor = connection.cursor(dictionary=True)

    # --------------------------------------------------------
    # TOTAL STUDENTS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'Student'
    """)

    students = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # TOTAL DONORS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'Donor'
    """)

    donors = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # TOTAL REQUESTS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_requests
    """)

    total_requests = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # PENDING REQUESTS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE status = 'Pending'
    """)

    pending_requests = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # COMPLETED REQUESTS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_requests
        WHERE status = 'Completed'
    """)

    completed_requests = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # TOTAL DONATIONS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_donations
    """)

    total_donations = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # AVAILABLE DONATIONS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_donations
        WHERE status = 'Available'
    """)

    available_donations = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # COLLECTED DONATIONS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM food_donations
        WHERE status = 'Collected'
    """)

    collected_donations = cursor.fetchone()["total"]

    # --------------------------------------------------------
    # REQUEST STATUS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            status,
            COUNT(*) AS total
        FROM food_requests
        GROUP BY status
    """)

    request_status = cursor.fetchall()

    # --------------------------------------------------------
    # DONATION STATUS
    # --------------------------------------------------------

    cursor.execute("""
        SELECT
            status,
            COUNT(*) AS total
        FROM food_donations
        GROUP BY status
    """)

    donation_status = cursor.fetchall()

    cursor.close()
    connection.close()

    # ========================================================
    # CREATE PDF
    # ========================================================

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]

    title_style.alignment = TA_CENTER

    heading_style = styles["Heading2"]

    normal_style = styles["BodyText"]

    elements = []

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "FoodLink Kimberley",
            title_style
        )
    )

    elements.append(
        Paragraph(
            "Administrative Report",
            heading_style
        )
    )

    elements.append(
        Paragraph(
            "Generated: "
            + datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            normal_style
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    # --------------------------------------------------------
    # SYSTEM OVERVIEW
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "System Overview",
            heading_style
        )
    )

    overview_data = [
        ["Metric", "Total"],
        ["Students", students],
        ["Donors", donors],
        ["Food Requests", total_requests],
        ["Pending Requests", pending_requests],
        ["Completed Requests", completed_requests],
        ["Food Donations", total_donations],
        ["Available Donations", available_donations],
        ["Collected Donations", collected_donations]
    ]

    overview_table = Table(
        overview_data,
        colWidths=[300, 100]
    )

    overview_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#064E3B")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                8
            )
        ])
    )

    elements.append(
        overview_table
    )

    elements.append(
        Spacer(1, 25)
    )

    # --------------------------------------------------------
    # REQUEST STATUS
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Food Requests by Status",
            heading_style
        )
    )

    request_table_data = [
        ["Status", "Total"]
    ]

    for item in request_status:

        request_table_data.append([
            item["status"],
            item["total"]
        ])

    request_table = Table(
        request_table_data,
        colWidths=[300, 100]
    )

    request_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#064E3B")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                8
            )
        ])
    )

    elements.append(
        request_table
    )

    elements.append(
        Spacer(1, 25)
    )

    # --------------------------------------------------------
    # DONATION STATUS
    # --------------------------------------------------------

    elements.append(
        Paragraph(
            "Food Donations by Status",
            heading_style
        )
    )

    donation_table_data = [
        ["Status", "Total"]
    ]

    for item in donation_status:

        donation_table_data.append([
            item["status"],
            item["total"]
        ])

    donation_table = Table(
        donation_table_data,
        colWidths=[300, 100]
    )

    donation_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#064E3B")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                8
            )
        ])
    )

    elements.append(
        donation_table
    )

    # --------------------------------------------------------
    # BUILD PDF
    # --------------------------------------------------------

    document.build(elements)

    output.seek(0)

    filename = (
        "FoodLink_Kimberley_Report_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".pdf"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# ============================================================
# LOGOUT
# ============================================================

@app.route("/logout")
def logout():

    # Store user before clearing session
    user_id = session.get("user_id")

    if user_id:

        log_activity(
            "Logout",
            "User logged out of FoodLink.",
            user_id
        )

    session.clear()

    return redirect(
        url_for("login")
    )


# ============================================================
# RUN APPLICATION
# ============================================================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
